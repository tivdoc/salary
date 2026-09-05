# L8-1 / D2: scripts refuse a production environment, before anything else.
import os as _tivdoc_os, sys as _tivdoc_sys
if _tivdoc_os.environ.get("NODE_ENV") == "production" or _tivdoc_os.environ.get("VERCEL_ENV") in ("production", "preview"):
    _tivdoc_sys.stderr.write("PRODUCTION_ENVIRONMENT_REFUSED\n")
    _tivdoc_sys.exit(2)

import csv
import io
import json
import sys

from openpyxl import load_workbook


MAX_SHEETS = 100
MAX_ROWS_PER_SHEET = 200_000
MAX_TEXT_CHARACTERS = 20_000_000


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def sheet_to_csv(worksheet) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    row_count = 0
    for row in worksheet.iter_rows(values_only=True):
        row_count += 1
        if row_count > MAX_ROWS_PER_SHEET:
            raise ValueError("xlsx_row_limit_exceeded")
        writer.writerow([cell_text(value) for value in row])
    return buffer.getvalue()


def main() -> int:
    if len(sys.argv) != 2:
        print(json.dumps({"safe_error_code": "xlsx_path_required"}))
        return 2
    try:
        # read_only avoids evaluating any embedded formulas/macros; data_only
        # takes the last-calculated cached value rather than a formula
        # string, keep_vba is left at its default (False) so a workbook that
        # relies on VBA is not silently treated as spreadsheet-only content.
        workbook = load_workbook(sys.argv[1], read_only=True, data_only=True, keep_vba=False)
    except Exception:
        print(json.dumps({"safe_error_code": "xlsx_parse_failed"}))
        return 4
    try:
        sheet_names = workbook.sheetnames
        if len(sheet_names) > MAX_SHEETS:
            print(json.dumps({"safe_error_code": "xlsx_sheet_limit_exceeded"}))
            return 5
        pages = []
        total_characters = 0
        for index, name in enumerate(sheet_names):
            worksheet = workbook[name]
            try:
                text = sheet_to_csv(worksheet)
            except ValueError as error:
                print(json.dumps({"safe_error_code": str(error)}))
                return 6
            total_characters += len(text)
            if total_characters > MAX_TEXT_CHARACTERS:
                print(json.dumps({"safe_error_code": "xlsx_text_limit_exceeded"}))
                return 7
            pages.append({"page": index + 1, "text": text, "sheet_name": name})
        print(json.dumps({"pages": pages}, ensure_ascii=False))
        return 0
    except Exception:
        print(json.dumps({"safe_error_code": "xlsx_parse_failed"}))
        return 4
    finally:
        workbook.close()


if __name__ == "__main__":
    raise SystemExit(main())
